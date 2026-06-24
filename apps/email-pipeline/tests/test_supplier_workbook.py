"""Characterization tests for supplier workbook parse, validation, and SQLite import."""

from __future__ import annotations

import ast
import inspect
import json
import sqlite3
from pathlib import Path
from zipfile import BadZipFile

import pytest
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException

import origenlab_email_pipeline.supplier_workbook as supplier_workbook_root
from origenlab_email_pipeline.core.suppliers import supplier_workbook as supplier_workbook_core
from origenlab_email_pipeline.supplier_schema import SUPPLIER_SCHEMA_SQL, ensure_supplier_tables
from origenlab_email_pipeline.supplier_workbook import (
    CONF_KEYS,
    CONTACT_ROUTE_KEYS,
    DOMAIN_KEYS,
    EXPECTED_SHEETS,
    FOCUS_KEYS,
    NAME_KEYS,
    RANK_KEYS,
    REGION_KEYS,
    TIER_ANEXO,
    TIER_EXCLUSION,
    TIER_TOP15,
    TIER_TOP50,
    build_company_domain_lookup,
    collect_workbook_validation_issues,
    extract_urls_from_row,
    import_supplier_workbook,
    infer_channel_type,
    load_workbook_sheets,
    merge_tier_labels,
    normalize_company_match_key,
    normalize_supplier_domain,
    parse_confidence,
    partition_supplier_validation_issues,
    pick_field,
    resolve_row_domain,
    sheet_as_dicts,
    sheet_as_dicts_skip_preamble,
    sheet_resumen_as_line_dicts,
    tier_weight,
)

_MODULE_PATH = Path(inspect.getfile(supplier_workbook_root))

_PUBLIC_CONSTANTS: tuple[str, ...] = (
    "EXPECTED_SHEETS",
    "TIER_TOP15",
    "TIER_TOP50",
    "TIER_ANEXO",
    "TIER_EXCLUSION",
    "DOMAIN_KEYS",
    "NAME_KEYS",
    "REGION_KEYS",
    "RANK_KEYS",
    "CONF_KEYS",
    "FOCUS_KEYS",
    "CONTACT_ROUTE_KEYS",
)

_PUBLIC_CALLABLES: tuple[str, ...] = (
    "tier_weight",
    "merge_tier_labels",
    "normalize_supplier_domain",
    "normalize_company_match_key",
    "sheet_as_dicts",
    "sheet_as_dicts_skip_preamble",
    "sheet_resumen_as_line_dicts",
    "pick_field",
    "parse_confidence",
    "extract_urls_from_row",
    "build_company_domain_lookup",
    "resolve_row_domain",
    "infer_channel_type",
    "load_workbook_sheets",
    "collect_workbook_validation_issues",
    "partition_supplier_validation_issues",
    "import_supplier_workbook",
)


def _conn() -> sqlite3.Connection:
    c = sqlite3.connect(":memory:")
    c.execute("PRAGMA foreign_keys=ON")
    return c


def _write_minimal_workbook(path: Path) -> None:
    wb = Workbook()
    r = wb.active
    r.title = "Resumen"
    r.append(["Campo", "Valor"])
    r.append(["Versión", "test-fixture"])

    op = wb.create_sheet("Oportunidades_50")
    op.append(["Dominio", "Empresa", "Región", "Ranking", "Confianza", "Foco"])
    op.append(["https://www.alpha.example.com", "Alpha", "Europa", 2, 0.85, "centrífugas"])
    op.append(["beta.example.com", "Beta", "Asia", 1, "90%", "microscopios"])

    c15 = wb.create_sheet("Contacto_15")
    c15.append(["Dominio", "Email contacto", "Formulario"])
    c15.append(["alpha.example.com", "sales@alpha.example.com", "https://alpha.example.com/contact"])

    ev = wb.create_sheet("Evidencias")
    ev.append(["Dominio", "URL", "Título"])
    ev.append(
        [
            "alpha.example.com",
            "https://catalog.alpha.example.com/listing",
            "Catálogo",
        ]
    )

    pr = wb.create_sheet("Prioridades")
    pr.append(["Categoría", "Prioridad"])
    pr.append(["centrífugas", 1])

    ex = wb.create_sheet("Exclusiones")
    ex.append(["Dominio", "Motivo"])
    ex.append(["legacy-old.com", "histórico"])

    ax = wb.create_sheet("Anexo_CSV_NoRepetido")
    ax.append(["Dominio", "Empresa", "Región", "Ranking"])
    ax.append(["gamma.example.org", "Gamma SA", "LATAM", 10])

    wb.save(path)


def _write_all_sheets_workbook(path: Path, *, op50_rows: list[list[object]] | None = None) -> None:
    wb = Workbook()
    wb.active.title = "Resumen"
    wb.active.append(["Campo", "Valor"])
    default_op = [["Dominio", "Empresa"], ["solo.example.com", "Solo Co"]]
    sheets: list[tuple[str, list[list[object]]]] = [
        ("Oportunidades_50", op50_rows or default_op),
        ("Contacto_15", [["Dominio"], ["solo.example.com"]]),
        ("Evidencias", [["Dominio", "URL"], ["solo.example.com", "https://solo.example.com/p"]]),
        ("Prioridades", [["c", "p"], ["x", 1]]),
        ("Exclusiones", [["Dominio"], []]),
        ("Anexo_CSV_NoRepetido", [["Dominio"], []]),
    ]
    for title, rows in sheets:
        ws = wb.create_sheet(title)
        for row in rows:
            ws.append(row)
    wb.save(path)


# --- 1. Public API / facade contract ---


def test_expected_sheets_contract() -> None:
    assert EXPECTED_SHEETS == frozenset(
        {
            "Resumen",
            "Oportunidades_50",
            "Contacto_15",
            "Evidencias",
            "Prioridades",
            "Exclusiones",
            "Anexo_CSV_NoRepetido",
        }
    )


@pytest.mark.parametrize("name", _PUBLIC_CONSTANTS)
def test_public_constants_on_root_and_core_facade(name: str) -> None:
    assert hasattr(supplier_workbook_root, name)
    assert hasattr(supplier_workbook_core, name)
    assert getattr(supplier_workbook_root, name) is getattr(supplier_workbook_core, name)


@pytest.mark.parametrize("name", _PUBLIC_CALLABLES)
def test_public_callables_same_reference_root_and_core(name: str) -> None:
    root_fn = getattr(supplier_workbook_root, name)
    core_fn = getattr(supplier_workbook_core, name)
    assert root_fn is core_fn
    assert callable(root_fn)


def test_tier_constants_and_weights() -> None:
    assert TIER_TOP15 == "top15"
    assert TIER_TOP50 == "top50"
    assert TIER_ANEXO == "anexo"
    assert TIER_EXCLUSION == "exclusion"
    assert tier_weight(TIER_TOP15) == 3
    assert tier_weight(TIER_TOP50) == 2
    assert tier_weight(TIER_ANEXO) == 1
    assert tier_weight(TIER_EXCLUSION) == 0
    assert tier_weight("unknown") == 0


def test_merge_tier_labels_precedence() -> None:
    assert merge_tier_labels(TIER_TOP50, TIER_TOP15) == TIER_TOP15
    assert merge_tier_labels(TIER_TOP15, TIER_TOP50) == TIER_TOP15
    assert merge_tier_labels(None, TIER_ANEXO) == TIER_ANEXO


# --- 2. Input parsing ---


def test_sheet_as_dicts_skips_blank_rows() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Dominio", "Empresa"])
    ws.append(["a.com", "A"])
    ws.append([None, None])
    ws.append(["", ""])
    rows = sheet_as_dicts(ws)
    assert len(rows) == 1
    assert rows[0]["dominio"] == "a.com"
    assert rows[0]["empresa"] == "A"


def test_sheet_as_dicts_english_header_aliases() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["website_domain", "company_name", "region_group", "source_rank", "confidence"])
    ws.append(["eng.example.com", "Eng Co", "EU", 3, "alta"])
    rows = sheet_as_dicts(ws)
    assert rows[0]["website_domain"] == "eng.example.com"
    assert pick_field(rows[0], DOMAIN_KEYS) == "eng.example.com"
    assert pick_field(rows[0], NAME_KEYS) == "Eng Co"
    assert pick_field(rows[0], REGION_KEYS) == "EU"
    assert pick_field(rows[0], RANK_KEYS) == 3


def test_sheet_as_dicts_skip_preamble_finds_anexo_header() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Anexo import"])
    ws.append([])
    ws.append(["website_domain", "company_name", "region_group", "source_rank"])
    ws.append(["anexo.example.com", "Anexo SA", "LATAM", 5])
    rows = sheet_as_dicts_skip_preamble(ws)
    assert len(rows) == 1
    assert rows[0]["company_name"] == "Anexo SA"


def test_sheet_resumen_as_line_dicts() -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Campo", "Valor"])
    ws.append(["Versión", "v1"])
    lines = sheet_resumen_as_line_dicts(ws)
    assert any("Campo" in str(r.get("line", "")) for r in lines)


def test_load_workbook_sheets_reports_missing_expected_sheets(tmp_path: Path) -> None:
    p = tmp_path / "partial.xlsx"
    wb = Workbook()
    wb.active.title = "Resumen"
    wb.active.append(["k", "v"])
    wb.save(p)
    sheets, issues = load_workbook_sheets(p)
    assert "Resumen" in sheets
    assert any(i.startswith("missing_sheet:") for i in issues)


def test_collect_validation_duplicate_domain_in_op50_sheet(tmp_path: Path) -> None:
    p = tmp_path / "dup_row.xlsx"
    _write_all_sheets_workbook(
        p,
        op50_rows=[
            ["Dominio", "Empresa"],
            ["x.com", "A"],
            ["x.com", "B"],
        ],
    )
    issues = collect_workbook_validation_issues(p)
    assert any("duplicate_domain_in_sheet_row:x.com" in x for x in issues)
    assert any(x == "oportunidades_50_duplicate_domain:x.com" for x in issues)


# --- 3. Normalization ---


def test_normalize_supplier_domain_strips_www_and_lower() -> None:
    assert normalize_supplier_domain("HTTPS://WWW.FOO.COM/path") == "foo.com"
    assert normalize_supplier_domain("delta.io") == "delta.io"
    assert normalize_supplier_domain("") is None
    assert normalize_supplier_domain(None) is None


def test_normalize_company_match_key_folds_punctuation() -> None:
    assert normalize_company_match_key("  Acme (Chile) S.A. ") == "acme chile s.a."
    assert normalize_company_match_key("") is None


def test_parse_confidence_numeric_and_word_labels() -> None:
    assert parse_confidence(0.85) == (0.85, "0.85")
    assert parse_confidence("90%")[0] == pytest.approx(0.9)
    score, label = parse_confidence("alta")
    assert score == pytest.approx(0.9)
    assert label == "alta"
    assert parse_confidence("not-a-score") == (None, "not-a-score")


def test_resolve_row_domain_via_company_lookup() -> None:
    op_rows = [{"company_name": "Beta Labs", "website_domain": "beta.example.com"}]
    lookup = build_company_domain_lookup(op_rows)
    contact_row = {"empresa": "Beta Labs"}
    assert resolve_row_domain(contact_row, lookup) == "beta.example.com"
    assert resolve_row_domain({"empresa": "Unknown"}, lookup) is None


def test_infer_channel_type() -> None:
    assert infer_channel_type("email contacto", "sales@x.com") == "email"
    assert infer_channel_type("formulario", "https://x.com/f") == "form"
    assert infer_channel_type("teléfono", "+56911112222") == "phone"
    assert infer_channel_type("linkedin", "https://www.linkedin.com/in/acme-lab") == "linkedin"
    assert infer_channel_type("linkedin", "https://uk.linkedin.com/company/acme") == "linkedin"
    assert infer_channel_type("linkedin", "https://linkedin.com.evil.test/profile") == "web"
    assert infer_channel_type("linkedin", "https://evil.test/?next=linkedin.com") == "web"
    assert infer_channel_type("linkedin", "not-a-url linkedin.com") == "other"


def test_extract_urls_from_row() -> None:
    row = {"url": "https://catalog.example.com/item (see also)"}
    urls = extract_urls_from_row(row)
    assert "https://catalog.example.com/item" in urls[0]


def test_pick_field_partial_key_match() -> None:
    row = {"email contacto": "ops@lab.cl"}
    assert pick_field(row, CONTACT_ROUTE_KEYS) == "ops@lab.cl"


# --- 4. SQLite / output contract ---


def test_supplier_schema_tables_in_ddl() -> None:
    for table in (
        "supplier_import_batch",
        "supplier_master",
        "supplier_evidence",
        "supplier_contact_channel",
        "supplier_priority_snapshot",
        "supplier_review_state",
    ):
        assert f"CREATE TABLE IF NOT EXISTS {table}" in SUPPLIER_SCHEMA_SQL


def test_import_creates_masters_evidence_snapshot(tmp_path: Path) -> None:
    p = tmp_path / "ok.xlsx"
    _write_minimal_workbook(p)
    conn = _conn()
    ensure_supplier_tables(conn)
    bid = import_supplier_workbook(conn, p)
    assert bid >= 1
    n_master = conn.execute("SELECT COUNT(*) FROM supplier_master").fetchone()[0]
    assert n_master >= 4
    alphas = conn.execute(
        "SELECT id FROM supplier_master WHERE domain_norm='alpha.example.com'"
    ).fetchone()
    assert alphas
    sid = int(alphas[0])
    tier = conn.execute(
        "SELECT tier FROM supplier_priority_snapshot WHERE supplier_id=? AND batch_id=?",
        (sid, bid),
    ).fetchone()[0]
    assert tier == TIER_TOP15
    n_ev = conn.execute(
        "SELECT COUNT(*) FROM supplier_evidence WHERE supplier_id=?", (sid,)
    ).fetchone()[0]
    assert n_ev >= 1
    ch = conn.execute(
        "SELECT channel_type FROM supplier_contact_channel WHERE supplier_id=? AND channel_type='email'",
        (sid,),
    ).fetchone()
    assert ch is not None


def test_import_batch_metadata_fields(tmp_path: Path) -> None:
    p = tmp_path / "meta.xlsx"
    _write_minimal_workbook(p)
    conn = _conn()
    ensure_supplier_tables(conn)
    bid = import_supplier_workbook(conn, p, source_filename="custom_name.xlsx")
    row = conn.execute(
        """
        SELECT source_filename, file_sha256, sheet_row_counts_json, category_priorities_json, resumen_note
        FROM supplier_import_batch WHERE id=?
        """,
        (bid,),
    ).fetchone()
    assert row[0] == "custom_name.xlsx"
    assert len(row[1]) == 64
    counts = json.loads(row[2])
    assert counts["Oportunidades_50"] >= 2
    assert counts["Resumen"] >= 1


def test_import_twice_creates_two_batches_one_master_per_domain(tmp_path: Path) -> None:
    p = tmp_path / "idem.xlsx"
    _write_all_sheets_workbook(p)
    conn = _conn()
    ensure_supplier_tables(conn)
    b1 = import_supplier_workbook(conn, p)
    b2 = import_supplier_workbook(conn, p)
    assert b2 > b1
    assert conn.execute("SELECT COUNT(*) FROM supplier_import_batch").fetchone()[0] == 2
    assert conn.execute("SELECT COUNT(*) FROM supplier_master").fetchone()[0] == 1
    domain = conn.execute("SELECT domain_norm FROM supplier_master").fetchone()[0]
    assert domain == "solo.example.com"


def test_annex_tier_recorded(tmp_path: Path) -> None:
    p = tmp_path / "ok.xlsx"
    _write_minimal_workbook(p)
    conn = _conn()
    ensure_supplier_tables(conn)
    bid = import_supplier_workbook(conn, p)
    row = conn.execute(
        "SELECT tier FROM supplier_priority_snapshot sm JOIN supplier_master m ON m.id=sm.supplier_id "
        "WHERE m.domain_norm='gamma.example.org' AND sm.batch_id=?",
        (bid,),
    ).fetchone()
    assert row and row[0] == TIER_ANEXO


def test_top50_snapshot_upgraded_by_contact15(tmp_path: Path) -> None:
    p = tmp_path / "ok.xlsx"
    _write_minimal_workbook(p)
    conn = _conn()
    ensure_supplier_tables(conn)
    bid = import_supplier_workbook(conn, p)
    row = conn.execute(
        "SELECT tier FROM supplier_priority_snapshot sm JOIN supplier_master m ON m.id=sm.supplier_id "
        "WHERE m.domain_norm='beta.example.com' AND sm.batch_id=?",
        (bid,),
    ).fetchone()
    assert row and row[0] == TIER_TOP50


def test_exclusion_keeps_flag_when_re_import_candidate(tmp_path: Path) -> None:
    p = tmp_path / "ex.xlsx"
    wb = Workbook()
    wb.active.title = "Resumen"
    wb.active.append(["a", "b"])
    for title, rows in [
        ("Oportunidades_50", [["Dominio", "Empresa"], ["both.com", "Both"]]),
        ("Contacto_15", [["Dominio"], ["both.com"]]),
        ("Evidencias", [["Dominio", "URL"], ["both.com", "https://u.example/x"]]),
        ("Prioridades", [["c"], [1]]),
        ("Exclusiones", [["Dominio"], ["both.com"]]),
        ("Anexo_CSV_NoRepetido", [["Dominio"], ["z99.com"]]),
    ]:
        ws = wb.create_sheet(title)
        for r in rows:
            ws.append(r)
    wb.save(p)
    conn = _conn()
    ensure_supplier_tables(conn)
    bid = import_supplier_workbook(conn, p)
    excl = conn.execute(
        "SELECT is_exclusion FROM supplier_master WHERE domain_norm='both.com'"
    ).fetchone()[0]
    assert int(excl) == 1
    tier = conn.execute(
        "SELECT tier FROM supplier_priority_snapshot sm "
        "JOIN supplier_master m ON m.id=sm.supplier_id "
        "WHERE m.domain_norm='both.com' AND sm.batch_id=?",
        (bid,),
    ).fetchone()[0]
    assert tier == TIER_EXCLUSION


# --- 5. Safety / import boundaries ---


def test_supplier_workbook_module_has_no_streamlit_imports() -> None:
    tree = ast.parse(_MODULE_PATH.read_text(encoding="utf-8"))
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                assert "streamlit" not in alias.name
        elif isinstance(node, ast.ImportFrom):
            mod = node.module or ""
            assert "streamlit" not in mod


def test_supplier_workbook_no_gmail_postgres_outbound_imports() -> None:
    text = _MODULE_PATH.read_text(encoding="utf-8").lower()
    forbidden = (
        "postgres",
        "psycopg",
        "sqlalchemy",
        "gmail_send",
        "send_email",
        "archive_send_batch",
        "archive_outreach",
        "outbound_core",
        "purge",
    )
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped.startswith(("import ", "from ")):
            continue
        lower = stripped.lower()
        for bad in forbidden:
            assert bad not in lower, f"unexpected import: {stripped}"


def test_module_docstring_supplier_workbook_focus() -> None:
    doc = supplier_workbook_root.__doc__ or ""
    assert "supplier" in doc.lower()
    assert "openpyxl" in doc.lower() or "workbook" in doc.lower()


# --- 6. Error behavior ---


def test_missing_workbook_file_raises_file_not_found(tmp_path: Path) -> None:
    missing = tmp_path / "no_such_workbook.xlsx"
    with pytest.raises(FileNotFoundError):
        load_workbook_sheets(missing)
    with pytest.raises(FileNotFoundError):
        collect_workbook_validation_issues(missing)


def test_corrupt_xlsx_raises_bad_zip_file(tmp_path: Path) -> None:
    bad = tmp_path / "corrupt.xlsx"
    bad.write_text("not a zip workbook", encoding="utf-8")
    with pytest.raises(BadZipFile):
        load_workbook_sheets(bad)


def test_csv_extension_raises_invalid_file_exception(tmp_path: Path) -> None:
    csv_path = tmp_path / "suppliers.csv"
    csv_path.write_text("domain,company\na.com,A\n", encoding="utf-8")
    with pytest.raises(InvalidFileException):
        load_workbook_sheets(csv_path)


def test_collect_validation_finds_annex_overlap(tmp_path: Path) -> None:
    p = tmp_path / "dup.xlsx"
    wb = Workbook()
    s = wb.active
    s.title = "Resumen"
    s.append(["k", "v"])
    op = wb.create_sheet("Oportunidades_50")
    op.append(["Dominio", "Empresa"])
    op.append(["a.com", "A"])
    c15 = wb.create_sheet("Contacto_15")
    c15.append(["Dominio"])
    c15.append(["a.com"])
    ev = wb.create_sheet("Evidencias")
    ev.append(["Dominio", "URL"])
    ev.append(["a.com", "https://x.com/e"])
    pr = wb.create_sheet("Prioridades")
    pr.append(["c", "p"])
    ex = wb.create_sheet("Exclusiones")
    ex.append(["Dominio"])
    ax = wb.create_sheet("Anexo_CSV_NoRepetido")
    ax.append(["Dominio"])
    ax.append(["a.com"])
    wb.save(p)

    issues = collect_workbook_validation_issues(p)
    assert any(x.startswith("anexo_overlap_op50:") for x in issues)


def test_partition_validation_errors_vs_warnings() -> None:
    err, warn = partition_supplier_validation_issues(
        ["missing_sheet:foo", "contacto_15_not_subset_op50:b.com", "evidencias_malformed_url:row_2:bad"]
    )
    assert any("missing_sheet" in x for x in err)
    assert any("evidencias_malformed_url" in x for x in err)
    assert any("contacto" in x for x in warn)
    assert not any("contacto" in x for x in err)
