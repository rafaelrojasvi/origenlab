"""Phase 2.4: selective content extraction for high-value attachment types.

This module is intentionally dependency-light and best-effort:
- No OCR (image-only PDFs remain empty).
- Always truncates stored text.
- Returns structured status/method + transparent heuristics.
"""

from __future__ import annotations

from dataclasses import dataclass
import csv as _csv
import io
import re
from typing import Literal

from origenlab_email_pipeline.timeutil import now_iso
import xml.etree.ElementTree as ET


ExtractStatus = Literal["success", "empty", "skipped", "unsupported", "failed"]
ExtractMethod = Literal["pdf_text", "docx", "xlsx", "csv", "xml", "none"]
DetectedDocType = Literal[
    "quote",
    "invoice",
    "price_list",
    "purchase_order",
    "datasheet",
    "unknown",
]


@dataclass(frozen=True)
class ExtractResult:
    status: ExtractStatus
    method: ExtractMethod
    text_preview: str
    text_truncated: str
    char_count: int
    page_count: int | None = None
    sheet_count: int | None = None
    detected_doc_type: DetectedDocType | None = None
    has_quote_terms: bool | None = None
    has_invoice_terms: bool | None = None
    has_price_list_terms: bool | None = None
    has_purchase_terms: bool | None = None
    error_message: str | None = None
    created_at: str | None = None


_RE_WS = re.compile(r"[ \t\r\f\v]+")


def _normalize_text(s: str) -> str:
    if not s:
        return ""
    s = s.replace("\x00", " ")
    s = _RE_WS.sub(" ", s)
    s = re.sub(r" *\n *", "\n", s)
    s = re.sub(r"\n\s*\n+", "\n\n", s)
    return s.strip()


def _truncate(s: str, *, preview_chars: int = 1600, max_chars: int = 50000) -> tuple[str, str]:
    s = _normalize_text(s)
    if not s:
        return "", ""
    prev = s[:preview_chars]
    trunc = s[:max_chars]
    return prev, trunc


def _lower(s: str | None) -> str:
    return (s or "").lower()


def guess_method(content_type: str | None, filename: str | None) -> ExtractMethod:
    ct = _lower(content_type)
    fn = _lower(filename)

    if ct.startswith("application/pdf") or fn.endswith(".pdf"):
        return "pdf_text"
    if fn.endswith(".docx") or "wordprocessingml.document" in ct:
        return "docx"
    if fn.endswith(".xlsx") or "spreadsheetml.sheet" in ct:
        return "xlsx"
    if fn.endswith(".csv") or ct == "text/csv":
        return "csv"
    if fn.endswith(".xml") or ct.endswith("/xml") or ct == "application/xml" or ct == "text/xml":
        return "xml"
    return "none"


def _signals_from_text(text: str, filename: str | None = None) -> tuple[DetectedDocType, dict[str, bool]]:
    t = _lower(text)
    fn = _lower(filename)
    blob = f"{t}\n{fn}"

    quote_terms = (
        "cotiz" in blob
        or "quotation" in blob
        or "quote" in blob
        or "presupuesto" in blob
        or "propuesta" in blob
    )
    invoice_terms = (
        "factura" in blob
        or "invoice" in blob
        or "boleta" in blob
        or "nota de crédito" in blob
        or "nota de credito" in blob
    )
    price_list_terms = (
        "lista de precios" in blob
        or "price list" in blob
        or "tarifario" in blob
        or "precio unit" in blob
        or "precio " in blob and "sku" in blob
    )
    purchase_terms = (
        "orden de compra" in blob
        or "purchase order" in blob
        or re.search(r"\boc\b", blob) is not None
        or "pedido" in blob
    )
    datasheet_terms = (
        "ficha técnica" in blob
        or "ficha tecnica" in blob
        or "datasheet" in blob
        or "specification" in blob
        or "technical data" in blob
    )

    # Conservative precedence: invoice > purchase > quote > price_list > datasheet
    if invoice_terms:
        doc_type: DetectedDocType = "invoice"
    elif purchase_terms:
        doc_type = "purchase_order"
    elif quote_terms:
        doc_type = "quote"
    elif price_list_terms:
        doc_type = "price_list"
    elif datasheet_terms:
        doc_type = "datasheet"
    else:
        doc_type = "unknown"

    return doc_type, {
        "has_quote_terms": quote_terms,
        "has_invoice_terms": invoice_terms,
        "has_price_list_terms": price_list_terms,
        "has_purchase_terms": purchase_terms,
    }


def extract_bytes(
    payload: bytes,
    *,
    content_type: str | None,
    filename: str | None,
    preview_chars: int = 1600,
    max_chars: int = 50000,
) -> ExtractResult:
    method = guess_method(content_type, filename)
    created_at = now_iso()

    if method == "none":
        return ExtractResult(
            status="unsupported",
            method="none",
            text_preview="",
            text_truncated="",
            char_count=0,
            detected_doc_type=None,
            created_at=created_at,
        )

    try:
        if method == "pdf_text":
            text, pages = _extract_pdf_text(payload)
            prev, trunc = _truncate(text, preview_chars=preview_chars, max_chars=max_chars)
            if not trunc:
                return ExtractResult(
                    status="empty",
                    method=method,
                    text_preview="",
                    text_truncated="",
                    char_count=0,
                    page_count=pages,
                    created_at=created_at,
                )
            dt, sig = _signals_from_text(trunc, filename)
            return ExtractResult(
                status="success",
                method=method,
                text_preview=prev,
                text_truncated=trunc,
                char_count=len(trunc),
                page_count=pages,
                detected_doc_type=dt,
                created_at=created_at,
                **sig,
            )

        if method == "docx":
            text = _extract_docx_text(payload)
            prev, trunc = _truncate(text, preview_chars=preview_chars, max_chars=max_chars)
            if not trunc:
                return ExtractResult(
                    status="empty",
                    method=method,
                    text_preview="",
                    text_truncated="",
                    char_count=0,
                    created_at=created_at,
                )
            dt, sig = _signals_from_text(trunc, filename)
            return ExtractResult(
                status="success",
                method=method,
                text_preview=prev,
                text_truncated=trunc,
                char_count=len(trunc),
                detected_doc_type=dt,
                created_at=created_at,
                **sig,
            )

        if method == "xlsx":
            text, sheet_count = _extract_xlsx_text(payload)
            prev, trunc = _truncate(text, preview_chars=preview_chars, max_chars=max_chars)
            if not trunc:
                return ExtractResult(
                    status="empty",
                    method=method,
                    text_preview="",
                    text_truncated="",
                    char_count=0,
                    sheet_count=sheet_count,
                    created_at=created_at,
                )
            dt, sig = _signals_from_text(trunc, filename)
            return ExtractResult(
                status="success",
                method=method,
                text_preview=prev,
                text_truncated=trunc,
                char_count=len(trunc),
                sheet_count=sheet_count,
                detected_doc_type=dt,
                created_at=created_at,
                **sig,
            )

        if method == "csv":
            text = _extract_csv_text(payload)
            prev, trunc = _truncate(text, preview_chars=preview_chars, max_chars=max_chars)
            if not trunc:
                return ExtractResult(
                    status="empty",
                    method=method,
                    text_preview="",
                    text_truncated="",
                    char_count=0,
                    created_at=created_at,
                )
            dt, sig = _signals_from_text(trunc, filename)
            return ExtractResult(
                status="success",
                method=method,
                text_preview=prev,
                text_truncated=trunc,
                char_count=len(trunc),
                detected_doc_type=dt,
                created_at=created_at,
                **sig,
            )

        if method == "xml":
            text = _extract_xml_text(payload)
            prev, trunc = _truncate(text, preview_chars=preview_chars, max_chars=max_chars)
            if not trunc:
                return ExtractResult(
                    status="empty",
                    method=method,
                    text_preview="",
                    text_truncated="",
                    char_count=0,
                    created_at=created_at,
                )
            dt, sig = _signals_from_text(trunc, filename)
            return ExtractResult(
                status="success",
                method=method,
                text_preview=prev,
                text_truncated=trunc,
                char_count=len(trunc),
                detected_doc_type=dt,
                created_at=created_at,
                **sig,
            )

        return ExtractResult(
            status="unsupported",
            method=method,
            text_preview="",
            text_truncated="",
            char_count=0,
            created_at=created_at,
        )
    except Exception as e:
        return ExtractResult(
            status="failed",
            method=method,
            text_preview="",
            text_truncated="",
            char_count=0,
            error_message=str(e)[:500],
            created_at=created_at,
        )


def _extract_pdf_text(payload: bytes) -> tuple[str, int | None]:
    # Lazy import so Phase 2.4 remains optional.
    import fitz  # type: ignore

    doc = fitz.open(stream=payload, filetype="pdf")
    pages = doc.page_count
    parts: list[str] = []
    for i in range(min(pages, 200)):  # hard cap pages for safety
        try:
            page = doc.load_page(i)
            parts.append(page.get_text("text") or "")
        except Exception:
            continue
    doc.close()
    return "\n\n".join(parts), pages


def _extract_docx_text(payload: bytes) -> str:
    import docx  # type: ignore

    bio = io.BytesIO(payload)
    d = docx.Document(bio)  # type: ignore[attr-defined]
    out: list[str] = []
    for p in d.paragraphs:
        if p.text and p.text.strip():
            out.append(p.text)
    return "\n".join(out)


def _extract_xlsx_text(payload: bytes) -> tuple[str, int | None]:
    from openpyxl import load_workbook  # type: ignore

    bio = io.BytesIO(payload)
    wb = load_workbook(bio, read_only=True, data_only=True)
    sheetnames = list(wb.sheetnames or [])
    out: list[str] = []
    out.append("SHEETS: " + ", ".join(sheetnames[:30]))
    # sample first 3 sheets
    for name in sheetnames[:3]:
        try:
            ws = wb[name]
        except Exception:
            continue
        out.append(f"\n[SHEET] {name}")
        # sample top-left grid: 30 rows x 12 cols
        r_cap, c_cap = 30, 12
        for r_i, row in enumerate(ws.iter_rows(max_row=r_cap, max_col=c_cap, values_only=True), start=1):
            if r_i > r_cap:
                break
            cells = []
            for v in row:
                if v is None:
                    cells.append("")
                else:
                    s = str(v)
                    s = s.replace("\n", " ").strip()
                    cells.append(s[:120])
            if any(cells):
                out.append(" | ".join(cells))
    try:
        wb.close()
    except Exception:
        pass
    return "\n".join(out), len(sheetnames)


def _decode_text_bytes(b: bytes) -> str:
    # Common encodings; fall back to utf-8 replacement
    for enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
        try:
            return b.decode(enc)
        except Exception:
            continue
    return b.decode("utf-8", errors="replace")


def _extract_csv_text(payload: bytes) -> str:
    s = _decode_text_bytes(payload)
    buf = io.StringIO(s)
    rdr = _csv.reader(buf)
    lines: list[str] = []
    max_rows = 200
    for i, row in enumerate(rdr):
        if i >= max_rows:
            break
        row = [c.replace("\n", " ").strip()[:200] for c in row]
        if any(row):
            lines.append(", ".join(row))
    return "\n".join(lines)


def _extract_xml_text(payload: bytes) -> str:
    s = _decode_text_bytes(payload)
    # Remove BOM/control chars that break parsing
    s = s.replace("\x00", "")
    root = ET.fromstring(s)  # may raise
    texts: list[str] = []
    # Collect text nodes up to a cap
    cap = 20000
    for elem in root.iter():
        if elem.text and elem.text.strip():
            t = elem.text.strip()
            texts.append(t)
            if sum(len(x) for x in texts) > cap:
                break
    return "\n".join(texts)

