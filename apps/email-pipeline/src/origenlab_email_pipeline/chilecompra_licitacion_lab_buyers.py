"""Heuristics to extract likely lab/clinical/QC-related buyer organizations from ChileCompra 'Licitación Publicada' exports."""

from __future__ import annotations

import csv
import re
from collections import defaultdict
from collections.abc import Iterable, Iterator, Mapping
from dataclasses import dataclass
from pathlib import Path
from typing import Any

# Line-item text: nombre adquisición, descripciones, UNSPSC labels.
_POSITIVE_PATTERNS: tuple[str, ...] = (
    r"laboratorio",
    r"\blab\b",
    r"hospital",
    r"cl[ií]nica",
    r"red\s*salud",
    r"servicio\s+(de\s+)?salud",
    r"universidad",
    r"instituto\s+de\s+salud\s+p[úu]blica",
    r"ispch",
    r"\binia\b",
    r"reactivo",
    r"diagn[oó]stic",
    r"microscop",
    r"centrífug",
    r"centrifug",
    r"electrofores",
    r"osm[oó]metr",
    r"ultrason",
    r"dispersor",
    r"hematolog",
    r"bioqu[ií]mic",
    r"microbiolog",
    r"patolog[ií]a",
    r"an[aá]lisis\s+(cl[ií]nico|qu[ií]mic)",
    r"control\s+de\s+calidad",
    r"inocuidad",
    r"calidad\s+aliment",
    r"equipo[s]?\s+(m[eé]dico|cl[ií]nico|de\s+laboratorio)",
    r"banco\s+de\s+sangre",
    r"tom(a|o)\s+de\s+muestras",
    r"biolog[ií]a\s+molecular",
    r"\bpcr\b",
    r"anatom[ií]a\s+patol",
    r"medicina\s+transfus",
    r"biotecnolog",
    r"facultad\s+de\s+(ciencias|qu[ií]mica|farmacia|medicina)",
    r"centro\s+de\s+servicios\s+anal[ií]ticos",
)

POSITIVE_RE = re.compile("|".join(f"({p})" for p in _POSITIVE_PATTERNS), re.IGNORECASE)

NEG_INFRA_ONLY_RE = re.compile(
    r"mejoramiento\s+de\s+calle|reparaci[oó]n\s+de\s+calzada|pavimentaci[oó]n\s+de\s+calle",
    re.IGNORECASE,
)

HEALTH_CONTEXT_RE = re.compile(
    r"hospital|cl[ií]nica|servicio\s+de\s+salud|salud\s+metropolitano|red\s*salud|ss(m|a)|minsal|de\s+salud\s+",
    re.IGNORECASE,
)


def classify_buyer_organization(name: str) -> str:
    """Coarse bucket for outreach triage (not a legal classification)."""
    n = name.lower()
    if re.search(r"complejo\s+asistencial|hhha|hbv|hospitaliz|\bhospital\b", n):
        return "hospital"
    if re.search(r"cl[ií]nica|clinicalascondes|christus|alemana|redsalud", n):
        return "clinica"
    if "universidad" in n or re.search(r"\b(uach|udec|uchile|uc\.cl|usach|utalca|ufro)\b", n):
        return "universidad"
    if re.search(r"servicio\s+de\s+salud|serv\.?\s*salud|red\s*salud|ssms|ssmoc|ssan\b", n):
        return "servicio_salud"
    if re.search(r"\binia\b|instituto\s+de\s+investigaci", n):
        return "instituto"
    if re.search(r"ispch|instituto\s+de\s+salud\s+p", n):
        return "isp"
    if re.search(r"municipalidad|i\.?\s*municipalidad", n):
        return "municipalidad"
    if re.search(r"corporaci[oó]n\s+municipal|corporacion\s+municipal", n):
        return "corporacion_municipal"
    return "otro"


def line_blob_from_row(row: Mapping[str, Any]) -> str:
    parts: list[str] = []
    for key in (
        "nombre_adquisicion",
        "descripcion",
        "descripcion_producto",
        "generico",
        "nivel_1",
        "nivel_2",
        "nivel_3",
    ):
        v = row.get(key)
        if v is not None and str(v).strip():
            parts.append(str(v))
    return " | ".join(parts)


def row_matches_lab_icp(*, organismo: str, blob: str) -> bool:
    if not POSITIVE_RE.search(blob):
        return False
    if NEG_INFRA_ONLY_RE.search(blob) and not HEALTH_CONTEXT_RE.search(
        organismo
    ) and not HEALTH_CONTEXT_RE.search(blob[:500]):
        return False
    return True


@dataclass(frozen=True)
class LicitacionLineRow:
    numero_adquisicion: str
    nombre_adquisicion: str
    descripcion: str
    organismo: str
    region: str
    descripcion_producto: str
    generico: str
    nivel_1: str
    nivel_2: str
    nivel_3: str

    def blob(self) -> str:
        return line_blob_from_row(
            {
                "nombre_adquisicion": self.nombre_adquisicion,
                "descripcion": self.descripcion,
                "descripcion_producto": self.descripcion_producto,
                "generico": self.generico,
                "nivel_1": self.nivel_1,
                "nivel_2": self.nivel_2,
                "nivel_3": self.nivel_3,
            }
        )


def aggregate_buyers_from_line_rows(
    rows: Iterable[LicitacionLineRow],
) -> list[dict[str, Any]]:
    """Aggregate unique (organismo, región) with counts and small samples."""
    by_key: dict[tuple[str, str], dict[str, Any]] = {}

    for r in rows:
        org = str(r.organismo or "").strip()
        region = str(r.region or "").strip()
        if not org:
            continue
        blob = r.blob()
        if not row_matches_lab_icp(organismo=org, blob=blob):
            continue
        key = (org, region)
        rec = by_key.get(key)
        if rec is None:
            rec = {
                "organization_name": org,
                "region": region,
                "buyer_type_guess": classify_buyer_organization(org),
                "matched_line_items": 0,
                "ids": set(),
                "titles": set(),
                "kw": defaultdict(int),
            }
            by_key[key] = rec
        rec["matched_line_items"] += 1
        num = str(r.numero_adquisicion or "").strip()
        if num and len(rec["ids"]) < 8:
            rec["ids"].add(num)
        nombre = str(r.nombre_adquisicion or "").strip()
        if nombre and len(rec["titles"]) < 5:
            rec["titles"].add(nombre[:240])
        for m in POSITIVE_RE.finditer(blob):
            rec["kw"][m.group(0).lower()[:50]] += 1

    out: list[dict[str, Any]] = []
    for rec in sorted(by_key.values(), key=lambda x: (-x["matched_line_items"], x["organization_name"])):
        top_kw = sorted(rec["kw"].items(), key=lambda kv: -kv[1])[:5]
        signal = "; ".join(k for k, _ in top_kw)
        out.append(
            {
                "organization_name": rec["organization_name"],
                "region": rec["region"],
                "buyer_type_guess": rec["buyer_type_guess"],
                "matched_line_items": rec["matched_line_items"],
                "sample_adquisicion_ids": ";".join(sorted(rec["ids"])),
                "sample_titles": " | ".join(list(rec["titles"])[:3]),
                "match_signal_keywords": signal[:500],
            }
        )
    return out


def iter_line_rows_from_chilecompra_xlsx(path: Path) -> Iterator[LicitacionLineRow]:
    """Read ChileCompra export 'Licitacion_Publicada.xlsx' (report layout as of 2026-04)."""
    from openpyxl import load_workbook  # lazy import

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        for row in ws.iter_rows(min_row=9, values_only=True):
            if not row or row[1] is None:
                continue
            cells = list(row)
            while len(cells) < 19:
                cells.append(None)

            yield LicitacionLineRow(
                numero_adquisicion=str(cells[1] or ""),
                nombre_adquisicion=str(cells[4] or ""),
                descripcion=str(cells[5] or ""),
                organismo=str(cells[6] or ""),
                region=str(cells[7] or ""),
                descripcion_producto=str(cells[11] or ""),
                generico=str(cells[15] or ""),
                nivel_1=str(cells[16] or ""),
                nivel_2=str(cells[17] or ""),
                nivel_3=str(cells[18] or ""),
            )
    finally:
        wb.close()


def write_buyers_csv(rows: list[dict[str, Any]], out_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "organization_name",
        "region",
        "buyer_type_guess",
        "matched_line_items",
        "sample_adquisicion_ids",
        "sample_titles",
        "match_signal_keywords",
    ]
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow({k: r[k] for k in fieldnames})


def extract_lab_buyers_to_csv(*, xlsx_path: Path, out_csv: Path) -> dict[str, int]:
    line_rows = iter_line_rows_from_chilecompra_xlsx(xlsx_path)
    agg = aggregate_buyers_from_line_rows(line_rows)
    write_buyers_csv(agg, out_csv)
    return {"buyer_rows": len(agg)}
