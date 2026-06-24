"""Parse / validate / import DeepSearch supplier workbooks (openpyxl).

Sheet contracts (v1):
  - Oportunidades_50, Contacto_15, Evidencias, Prioridades, Exclusiones, Anexo_CSV_NoRepetido, Resumen
"""

from __future__ import annotations

import hashlib
import json
import re
import sqlite3
from collections import Counter
from collections.abc import Iterable, Mapping
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import load_workbook

from origenlab_email_pipeline.timeutil import now_iso
from origenlab_email_pipeline.supplier_schema import ensure_supplier_tables

# ---------------------------------------------------------------------------
# 1. Constants / sheet names / tier names / column aliases
# ---------------------------------------------------------------------------

EXPECTED_SHEETS = frozenset(
    {
        "Resumen",
        "Oportunidades_50",
        "Contacto_15",
        "Evidencias",
        "Prioridades",
        "Exclusiones",
        "Anexo_CSV_NoRepetido",
    }
)

TIER_TOP15 = "top15"
TIER_TOP50 = "top50"
TIER_ANEXO = "anexo"
TIER_EXCLUSION = "exclusion"

DOMAIN_KEYS = (
    "dominio",
    "domain",
    "sitio web",
    "website",
    "website_domain",
    "supplier_domain",
    "url web",
    "sitio",
    "host",
)
NAME_KEYS = (
    "empresa",
    "company_name",
    "proveedor",
    "organización",
    "organizacion",
    "nombre",
    "nombre comercial",
    "company",
    "razón social",
    "razon social",
    "fabricante",
)
REGION_KEYS = (
    "región",
    "region",
    "region_group",
    "region_guess",
    "país",
    "pais",
    "country",
    "mercado",
    "zona",
)
RANK_KEYS = (
    "ranking",
    "rank",
    "source_rank",
    "posición",
    "posicion",
    "#",
    "orden",
    "n°",
    "no.",
    "número",
    "numero",
    "prioridad",
    "puesto",
)
CONF_KEYS = ("confianza", "confidence", "score", "puntaje", "nivel", "fiabilidad")
FOCUS_KEYS = (
    "foco",
    "categoría",
    "categoria",
    "equipos",
    "equipment",
    "equipment_focus",
    "equipo_foco",
    "tags",
    "línea",
    "linea",
    "vertical",
)

CONTACT_ROUTE_KEYS = (
    "contact_channel",
    "canal_contacto",
    "canal contacto",
    "email contacto",
    "contacto",
)

_URL_RE = re.compile(r"https?://[^\s]+", re.IGNORECASE)

_SHEET_RESUMEN = "Resumen"
_SHEET_ANEXO = "Anexo_CSV_NoRepetido"
_SHEET_OP50 = "Oportunidades_50"
_SHEET_CONTACTO15 = "Contacto_15"
_SHEET_EVIDENCIAS = "Evidencias"
_SHEET_EXCLUSIONES = "Exclusiones"

_EMAIL_COLUMN_TOKENS = ("email", "mail", "correo", "contact")


# ---------------------------------------------------------------------------
# 2. Tier ranking helpers
# ---------------------------------------------------------------------------


def tier_weight(tier: str | None) -> int:
    return {
        TIER_TOP15: 3,
        TIER_TOP50: 2,
        TIER_ANEXO: 1,
        TIER_EXCLUSION: 0,
    }.get((tier or "").strip().lower(), 0)


def merge_tier_labels(current: str | None, incoming: str) -> str:
    if tier_weight(incoming) > tier_weight(current):
        return incoming
    return current or incoming


# ---------------------------------------------------------------------------
# 3. Generic parsing helpers
# ---------------------------------------------------------------------------


def _norm_header_key(h: object) -> str:
    t = str(h).strip().lower() if h is not None else ""
    t = re.sub(r"\s+", " ", t)
    return t


def _rows_to_dicts(headers: list[str], data_rows: Iterable[tuple[Any, ...] | None]) -> list[dict[str, object | None]]:
    out: list[dict[str, object | None]] = []
    for row in data_rows:
        if row is None:
            continue
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        d: dict[str, object | None] = {}
        for i, key in enumerate(headers):
            if not key:
                continue
            d[key] = row[i] if i < len(row) else None
        if d:
            out.append(d)
    return out


def sheet_as_dicts(ws) -> list[dict[str, object | None]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_norm_header_key(h) for h in rows[0]]
    return _rows_to_dicts(headers, rows[1:])


def sheet_as_dicts_skip_preamble(ws) -> list[dict[str, object | None]]:
    """Find a true header row (e.g. Anexo with title blocks above the table)."""
    rows = list(ws.iter_rows(values_only=True))
    start_idx = -1
    for i, row in enumerate(rows):
        flat: set[str] = set()
        for c in row:
            if c is None or str(c).strip() == "":
                continue
            flat.add(_norm_header_key(c))
        if "website_domain" in flat and ("company_name" in flat or "source_rank" in flat):
            start_idx = i
            break
    if start_idx < 0:
        return sheet_as_dicts(ws)
    hdr = rows[start_idx]
    headers = [_norm_header_key(h) for h in hdr]
    return _rows_to_dicts(headers, rows[start_idx + 1 :])


def sheet_resumen_as_line_dicts(ws) -> list[dict[str, object | None]]:
    """Resumen is a narrative grid, not a keyed table — one pseudo-column per non-empty row."""
    rows = list(ws.iter_rows(values_only=True))[:60]
    out: list[dict[str, object | None]] = []
    for row in rows:
        parts = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if parts:
            out.append({"line": " | ".join(parts)})
    return out


def pick_field(row: Mapping[str, object | None], aliases: Iterable[str]) -> object | None:
    keys = list(row.keys())
    for a in aliases:
        if a in row:
            v = row.get(a)
            if v is not None and str(v).strip():
                return v
    for a in aliases:
        for k in keys:
            if a == k or (a in k and len(a) >= 3):
                v = row.get(k)
                if v is not None and str(v).strip():
                    return v
    return None


def parse_confidence(val: object | None) -> tuple[float | None, str | None]:
    if val is None:
        return None, None
    label = str(val).strip()
    if not label:
        return None, None
    try:
        x = float(label.replace(",", ".").replace("%", "").strip())
        if "%" in str(val):
            x = x / 100.0 if x > 1.0 else x
        if x > 1.0 and x <= 100.0:
            x = x / 100.0
        if x > 1.0 or x < 0.0:
            return None, label
        return x, label
    except ValueError:
        pass
    low = label.lower()
    word_scores = {
        "high": 0.9,
        "alta": 0.9,
        "medium": 0.65,
        "media": 0.65,
        "low": 0.35,
        "baja": 0.35,
        "muy alta": 0.95,
        "muy baja": 0.2,
    }
    if low in word_scores:
        return word_scores[low], label
    return None, label


def _ints_from_row(row: dict[str, object | None], aliases: Iterable[str]) -> int | None:
    v = pick_field(row, aliases)
    if v is None:
        return None
    try:
        return int(float(str(v).replace(",", ".")))
    except ValueError:
        return None


def extract_urls_from_row(row: Mapping[str, object | None]) -> list[str]:
    urls: list[str] = []
    for k, v in row.items():
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        kl = str(k).lower()
        if s.startswith("http") or "http" in kl or "url" in kl or "enlace" in kl or "link" in kl or "evidencia" in kl:
            for m in _URL_RE.findall(s):
                u = m.rstrip(").,;]")
                if u not in urls:
                    urls.append(u)
        elif s.startswith("www."):
            urls.append("https://" + s.split()[0].rstrip(").,;]"))
    cell_url = pick_field(row, ("evidence_url", "url evidencia", "enlace", "link"))
    if cell_url is not None:
        for m in _URL_RE.findall(str(cell_url).strip()):
            found = m.rstrip(").,;]")
            if found not in urls:
                urls.append(found)
    return urls


def _is_plausible_url(u: str) -> bool:
    try:
        p = urlparse(u)
        return p.scheme in ("http", "https") and bool(p.netloc)
    except ValueError:
        return False


def _normalize_channel_value(channel_type: str, raw: str) -> str:
    t = channel_type.lower()
    s = raw.strip()
    if t == "email":
        return s.lower()
    return s


def _normalize_url_hostname(host: str) -> str:
    normalized = host.strip().lower().split(":", 1)[0]
    if normalized.startswith("www."):
        normalized = normalized[4:]
    return normalized


def _hostname_matches_domain(host: str, domain: str) -> bool:
    normalized = _normalize_url_hostname(host)
    base = domain.strip().lower()
    if not normalized or not base:
        return False
    return normalized == base or normalized.endswith(f".{base}")


def _url_hostname(value: str) -> str | None:
    raw = value.strip()
    if not raw or raw.lower().startswith("//"):
        return None
    if not re.match(r"https?://", raw, re.I):
        return None
    try:
        parsed = urlparse(raw)
    except ValueError:
        return None
    if parsed.scheme.lower() not in ("http", "https") or not parsed.hostname:
        return None
    return parsed.hostname.lower()


def infer_channel_type(key: str, value: str) -> str:
    kl = key.lower()
    v = value.lower()
    if "@" in value:
        return "email"
    linkedin_host = _url_hostname(value)
    if linkedin_host and _hostname_matches_domain(linkedin_host, "linkedin.com"):
        return "linkedin"
    if "form" in kl or "formulario" in kl:
        return "form"
    if "tel" in kl or "phone" in kl or "telefono" in kl or "teléfono" in kl:
        return "phone"
    if value.startswith("http"):
        return "web"
    return "other"


# ---------------------------------------------------------------------------
# 4. Workbook loading and sheet extraction
# ---------------------------------------------------------------------------


def _parse_workbook_sheet(ws, sheet_name: str) -> list[dict[str, object | None]]:
    if sheet_name == _SHEET_RESUMEN:
        return sheet_resumen_as_line_dicts(ws)
    if sheet_name == _SHEET_ANEXO:
        return sheet_as_dicts_skip_preamble(ws)
    return sheet_as_dicts(ws)


def load_workbook_sheets(path: Path) -> tuple[dict[str, list[dict[str, object | None]]], list[str]]:
    """Return sheet dicts and list of issues (missing expected sheets)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        found = frozenset(wb.sheetnames)
        issues = [f"missing_sheet:{s}" for s in sorted(EXPECTED_SHEETS - found)]
        out: dict[str, list[dict[str, object | None]]] = {}
        for name in wb.sheetnames:
            out[name] = _parse_workbook_sheet(wb[name], name)
        return out, issues
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# 5. Domain / company normalization
# ---------------------------------------------------------------------------


def normalize_supplier_domain(raw: object | None) -> str | None:
    """Lowercase host, strip www., tolerate full URLs."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    s = s.lower().replace("\\", "/")
    if "://" not in s and "/" in s:
        s = "https://" + s.split("/")[0]
    if "://" in s:
        parsed = urlparse(s if "://" in s else f"https://{s}")
        host = (parsed.netloc or parsed.path or "").strip().lower()
    else:
        host = s.split("/")[0].strip().lower()
    host = host.split(":")[0]
    if host.startswith("www."):
        host = host[4:]
    return host or None


def normalize_company_match_key(raw: object | None) -> str | None:
    """Fold company labels for matching Contacto_15 (empresa) ↔ Oportunidades (company_name)."""
    if raw is None:
        return None
    s = str(raw).strip().lower()
    if not s:
        return None
    s = re.sub(r"[\(\)]", " ", s)
    s = re.sub(r"[^\w\s&\-.]", " ", s, flags=re.UNICODE)
    s = re.sub(r"\s+", " ", s).strip()
    return s or None


def build_company_domain_lookup(rows: list[dict[str, object | None]]) -> dict[str, str]:
    """Map normalized company name → domain from the main opportunity sheet."""
    out: dict[str, str] = {}
    for row in rows:
        d = normalize_supplier_domain(pick_field(row, DOMAIN_KEYS))
        if not d:
            continue
        nm = pick_field(row, NAME_KEYS)
        if nm is None:
            continue
        k = normalize_company_match_key(nm)
        if k:
            out[k] = d
    return out


def resolve_row_domain(
    row: Mapping[str, object | None],
    company_lookup: dict[str, str] | None,
) -> str | None:
    d_raw = pick_field(row, DOMAIN_KEYS)
    dn = normalize_supplier_domain(d_raw)
    if dn:
        return dn
    if not company_lookup:
        return None
    emp = pick_field(row, NAME_KEYS)
    if emp is None:
        return None
    k = normalize_company_match_key(emp)
    if not k:
        return None
    return company_lookup.get(k)


# ---------------------------------------------------------------------------
# 6. Validation / issue partitioning
# ---------------------------------------------------------------------------


def _domains_from_sheet_rows(
    rows: list[dict[str, object | None]],
    company_lookup: dict[str, str],
    *,
    start_row: int = 2,
) -> tuple[list[str], list[str]]:
    doms: list[str] = []
    bad: list[str] = []
    seen_row_dup: set[str] = set()
    for i, row in enumerate(rows, start=start_row):
        dn = resolve_row_domain(row, company_lookup)
        if not dn:
            if pick_field(row, NAME_KEYS) or pick_field(row, RANK_KEYS):
                bad.append(f"row_{i}_missing_domain")
            continue
        doms.append(dn)
        if dn in seen_row_dup:
            bad.append(f"duplicate_domain_in_sheet_row:{dn}")
        seen_row_dup.add(dn)
    return doms, bad


def collect_workbook_validation_issues(path: Path) -> list[str]:
    """Structural checks without importing (duplicate domains, annex overlap, orphan evidence)."""
    sheets, missing = load_workbook_sheets(path)
    issues = list(missing)

    op50 = sheets.get(_SHEET_OP50) or []
    c15 = sheets.get(_SHEET_CONTACTO15) or []
    anexo = sheets.get(_SHEET_ANEXO) or []
    excl = sheets.get(_SHEET_EXCLUSIONES) or []
    evid = sheets.get(_SHEET_EVIDENCIAS) or []
    company_lookup = build_company_domain_lookup(op50)

    op_list, op_bad = _domains_from_sheet_rows(op50, company_lookup)
    issues.extend(f"oportunidades_50:{x}" for x in op_bad)
    if len(op_list) > 55:
        issues.append(f"oportunidades_50_row_count_sanity:{len(op_list)}")
    op_counts = Counter(op_list)
    for d, n in op_counts.items():
        if n > 1:
            issues.append(f"oportunidades_50_duplicate_domain:{d}")
            break

    c_list, c_bad = _domains_from_sheet_rows(c15, company_lookup)
    issues.extend(f"contacto_15:{x}" for x in c_bad)
    cset, opset = set(c_list), set(op_list)
    if cset - opset:
        issues.append(
            "contacto_15_not_subset_op50:" + ",".join(sorted(cset - opset)[:12])
            + ("…" if len(cset - opset) > 12 else "")
        )

    ax_list, ax_bad = _domains_from_sheet_rows(anexo, company_lookup)
    issues.extend(f"anexo:{x}" for x in ax_bad)
    overlap = set(ax_list) & opset
    if overlap:
        issues.append(
            "anexo_overlap_op50:" + ",".join(sorted(overlap)[:15]) + ("…" if len(overlap) > 15 else "")
        )

    ex_list, ex_bad = _domains_from_sheet_rows(excl, company_lookup)
    issues.extend(f"exclusiones:{x}" for x in ex_bad)

    ev_domains: list[str] = []
    for i, row in enumerate(evid, start=2):
        dn = resolve_row_domain(row, company_lookup)
        urls = extract_urls_from_row(row)
        if not urls:
            u = pick_field(row, ("url", "enlace", "link", "evidencia url", "fuente"))
            if u is not None:
                urls = extract_urls_from_row({"_": u})
        for u in urls:
            if not _is_plausible_url(u):
                issues.append(f"evidencias_malformed_url:row_{i}:{u[:80]}")
        if dn:
            ev_domains.append(dn)

    ev_set = set(ev_domains)
    known = opset | cset | set(ax_list) | set(ex_list)
    orphans = ev_set - known
    if orphans and ev_set:
        issues.append(
            "evidence_domain_unknown:" + ",".join(sorted(orphans)[:12])
            + ("…" if len(orphans) > 12 else "")
        )

    missing_ranks = 0
    for row in op50:
        if resolve_row_domain(row, company_lookup) and _ints_from_row(row, RANK_KEYS) is None:
            missing_ranks += 1
    if missing_ranks:
        issues.append(f"oportunidades_50_missing_rank_count:{missing_ranks}")

    return issues


def partition_supplier_validation_issues(
    issues: list[str],
) -> tuple[list[str], list[str]]:
    """Split into (errors, warnings) for CI / scripts."""

    def _is_err(msg: str) -> bool:
        return bool(
            msg.startswith("missing_sheet:")
            or msg.startswith("evidencias_malformed_url:")
            or msg.startswith("evidence_domain_unknown:")
        )

    errors: list[str] = []
    warnings: list[str] = []
    for msg in issues:
        (errors if _is_err(msg) else warnings).append(msg)
    return errors, warnings


# ---------------------------------------------------------------------------
# 7. Summary / batch note helpers
# ---------------------------------------------------------------------------


def _read_docx_plain_text(path: Path, *, max_chars: int = 14_000) -> str | None:
    try:
        from docx import Document  # type: ignore[import-untyped]
    except ImportError:
        return None
    doc = Document(str(path))
    paras = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    blob = "\n".join(paras)
    return blob[:max_chars] if blob else None


def _build_batch_resumen_note(
    res_rows: list[dict[str, object | None]],
    informe_docx: Path | None,
) -> str:
    resumen_text = ""
    if res_rows:
        bits = []
        for r in res_rows[:30]:
            bits.append(" | ".join(str(v) for v in r.values() if v is not None and str(v).strip()))
        resumen_text = "\n".join(bits)[:8000]
    if informe_docx is not None and informe_docx.is_file():
        doc_txt = _read_docx_plain_text(informe_docx.resolve())
        if doc_txt:
            resumen_text = (
                (resumen_text + "\n\n--- Informe Word (referencia) ---\n\n" + doc_txt)[:20_000]
                if resumen_text
                else doc_txt[:20_000]
            )
    return resumen_text


def _insert_import_batch(
    conn: sqlite3.Connection,
    *,
    source_filename: str,
    file_sha256: str,
    imported_at: str,
    row_counts: dict[str, int],
    category_priorities_json: str,
    resumen_note: str,
) -> int:
    cur = conn.execute(
        """
        INSERT INTO supplier_import_batch (
          source_filename, file_sha256, imported_at,
          sheet_row_counts_json, category_priorities_json, resumen_note
        ) VALUES (?, ?, ?, ?, ?, ?)
        """,
        (
            source_filename,
            file_sha256,
            imported_at,
            json.dumps(row_counts, ensure_ascii=False),
            category_priorities_json,
            resumen_note or None,
        ),
    )
    return int(cur.lastrowid)


# ---------------------------------------------------------------------------
# 8. SQLite import / upsert logic
# ---------------------------------------------------------------------------


def _upsert_supplier_master(
    conn: sqlite3.Connection,
    domain_norm: str,
    now: str,
    *,
    trade_name: str | None,
    website: str | None,
    region: str | None,
    country: str | None,
    equipment: str | None,
    notes: str | None,
    is_exclusion: bool,
) -> int:
    trade_name = (trade_name or "").strip() or None
    website = (website or "").strip() or None
    region = (region or "").strip() or None
    country = (country or "").strip() or None
    equipment = (equipment or "").strip() or None
    notes = (notes or "").strip() or None
    row = conn.execute(
        "SELECT id, trade_name, website, region_label, country_label, equipment_focus, is_exclusion FROM supplier_master WHERE domain_norm=?",
        (domain_norm,),
    ).fetchone()
    if row is None:
        icur = conn.execute(
            """
            INSERT INTO supplier_master (
              domain_norm, trade_name, website, region_label, country_label, equipment_focus,
              notes, is_exclusion, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                domain_norm,
                trade_name,
                website,
                region,
                country,
                equipment,
                notes,
                1 if is_exclusion else 0,
                now,
                now,
            ),
        )
        sid = int(icur.lastrowid)
    else:
        sid = int(row[0])
        prev_excl = int(row[6] or 0)
        excl = 1 if (is_exclusion or prev_excl) else 0
        conn.execute(
            """
            UPDATE supplier_master SET
              trade_name = COALESCE(?, trade_name),
              website = COALESCE(?, website),
              region_label = COALESCE(?, region_label),
              country_label = COALESCE(?, country_label),
              equipment_focus = COALESCE(?, equipment_focus),
              notes = COALESCE(?, notes),
              is_exclusion = ?,
              updated_at = ?
            WHERE id = ?
            """,
            (
                trade_name,
                website,
                region,
                country,
                equipment,
                notes,
                excl,
                now,
                sid,
            ),
        )
    conn.execute(
        "INSERT OR IGNORE INTO supplier_review_state (supplier_id, status) VALUES (?, 'nuevo')",
        (sid,),
    )
    return sid


def _apply_priority_snapshot(
    conn: sqlite3.Connection,
    supplier_id: int,
    batch_id: int,
    tier: str,
    rank: int | None,
    conf_score: float | None,
    conf_label: str | None,
    cat_ctx: str | None,
) -> None:
    cur_snap = conn.execute(
        """
        SELECT tier, rank_in_list, confidence_score, confidence_label, category_context
        FROM supplier_priority_snapshot WHERE supplier_id=? AND batch_id=?
        """,
        (supplier_id, batch_id),
    ).fetchone()
    if cur_snap is None:
        conn.execute(
            """
            INSERT INTO supplier_priority_snapshot (
              supplier_id, batch_id, tier, rank_in_list,
              confidence_score, confidence_label, category_context
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (supplier_id, batch_id, tier, rank, conf_score, conf_label, cat_ctx),
        )
        return
    ot, orank, oconf, olabel, occ = cur_snap
    nt = merge_tier_labels(ot, tier)
    nr = orank
    nconf = oconf
    nlabel = olabel
    ncc = occ or cat_ctx
    if tier_weight(tier) > tier_weight(ot):
        nr = rank
        nconf = conf_score if conf_score is not None else oconf
        nlabel = conf_label or olabel
        ncc = cat_ctx or occ
    elif tier_weight(tier) == tier_weight(ot):
        ranks = [x for x in (orank, rank) if x is not None]
        nr = min(ranks) if ranks else orank
        if conf_score is not None and (oconf is None or conf_score > oconf):
            nconf = conf_score
            nlabel = conf_label or olabel
        if cat_ctx:
            ncc = cat_ctx
    conn.execute(
        """
        UPDATE supplier_priority_snapshot SET
          tier = ?, rank_in_list = ?, confidence_score = ?, confidence_label = ?, category_context = ?
        WHERE supplier_id = ? AND batch_id = ?
        """,
        (nt, nr, nconf, nlabel, ncc, supplier_id, batch_id),
    )


def _insert_row_evidence_urls(
    conn: sqlite3.Connection,
    supplier_id: int,
    batch_id: int,
    row: Mapping[str, object | None],
    source_sheet: str,
) -> None:
    for u in extract_urls_from_row(row):
        if _is_plausible_url(u):
            conn.execute(
                """
                INSERT OR IGNORE INTO supplier_evidence (supplier_id, batch_id, url, source_sheet)
                VALUES (?, ?, ?, ?)
                """,
                (supplier_id, batch_id, u, source_sheet),
            )


def _insert_row_contact_channels(
    conn: sqlite3.Connection,
    supplier_id: int,
    batch_id: int,
    row: Mapping[str, object | None],
    tier: str,
    source_sheet: str,
) -> None:
    route = pick_field(row, CONTACT_ROUTE_KEYS)
    if route and str(route).strip():
        rt = str(route).strip()
        pref_rt = 1 if tier == TIER_TOP15 else 0
        conn.execute(
            """
            INSERT OR IGNORE INTO supplier_contact_channel (
              supplier_id, batch_id, channel_type, value_raw, value_normalized, is_preferred, source_sheet
            ) VALUES (?, ?, 'contact_route', ?, ?, ?, ?)
            """,
            (supplier_id, batch_id, rt, rt[:512], pref_rt, source_sheet),
        )
    for k, v in row.items():
        if v is None or "@" not in str(v):
            continue
        kl = str(k).lower()
        if not any(tok in kl for tok in _EMAIL_COLUMN_TOKENS):
            continue
        for part in re.split(r"[\s,;]+", str(v)):
            part = part.strip("<>()[]")
            if "@" in part and "." in part:
                em = part
                pref_em = 1 if tier == TIER_TOP15 else 0
                conn.execute(
                    """
                    INSERT OR IGNORE INTO supplier_contact_channel (
                      supplier_id, batch_id, channel_type, value_raw, value_normalized, is_preferred, source_sheet
                    ) VALUES (?, ?, 'email', ?, ?, ?, ?)
                    """,
                    (supplier_id, batch_id, em, em.lower(), pref_em, source_sheet),
                )


def _ingest_candidate_rows(
    conn: sqlite3.Connection,
    batch_id: int,
    company_lookup: dict[str, str],
    now: str,
    rows: list[dict[str, object | None]],
    tier: str,
    source_sheet: str,
) -> None:
    for row in rows:
        dn = resolve_row_domain(row, company_lookup)
        if not dn:
            continue
        name = pick_field(row, NAME_KEYS)
        name_s = str(name).strip() if name is not None else None
        region = pick_field(row, REGION_KEYS)
        reg_s = str(region).strip() if region is not None else None
        country_v = pick_field(row, ("country", "país", "pais"))
        country_s = str(country_v).strip() if country_v is not None else None
        website = None
        wv = pick_field(row, ("website", "sitio web", "url"))
        if wv is not None:
            website = str(wv).strip()
        if not website and dn:
            website = f"https://{dn}"
        equip = pick_field(row, FOCUS_KEYS)
        equip_s = str(equip).strip() if equip is not None else None
        rk = _ints_from_row(row, RANK_KEYS)
        cs, cl = parse_confidence(pick_field(row, CONF_KEYS))
        sid = _upsert_supplier_master(
            conn,
            dn,
            now,
            trade_name=name_s,
            website=website,
            region=reg_s,
            country=country_s,
            equipment=equip_s,
            notes=None,
            is_exclusion=False,
        )
        _apply_priority_snapshot(conn, sid, batch_id, tier, rk, cs, cl, equip_s)
        _insert_row_evidence_urls(conn, sid, batch_id, row, source_sheet)
        _insert_row_contact_channels(conn, sid, batch_id, row, tier, source_sheet)


def _ingest_exclusion_rows(
    conn: sqlite3.Connection,
    batch_id: int,
    company_lookup: dict[str, str],
    now: str,
    rows: list[dict[str, object | None]],
) -> None:
    for row in rows:
        dn = resolve_row_domain(row, company_lookup)
        if not dn:
            continue
        note = pick_field(row, ("motivo", "nota", "razón", "razon", "comentario", "notes"))
        note_s = str(note).strip() if note is not None else None
        sid = _upsert_supplier_master(
            conn,
            dn,
            now,
            trade_name=str(pick_field(row, NAME_KEYS) or "").strip() or None,
            website=None,
            region=str(pick_field(row, REGION_KEYS) or "").strip() or None,
            country=None,
            equipment=None,
            notes=note_s,
            is_exclusion=True,
        )
        _apply_priority_snapshot(conn, sid, batch_id, TIER_EXCLUSION, None, None, "exclusión", note_s)


def _ingest_evidence_rows(
    conn: sqlite3.Connection,
    batch_id: int,
    company_lookup: dict[str, str],
    rows: list[dict[str, object | None]],
) -> None:
    for row in rows:
        dn = resolve_row_domain(row, company_lookup)
        if not dn:
            continue
        sid_row = conn.execute(
            "SELECT id FROM supplier_master WHERE domain_norm=?", (dn,)
        ).fetchone()
        if sid_row is None:
            continue
        sid = int(sid_row[0])
        tit = pick_field(row, ("título", "titulo", "title"))
        snip = pick_field(row, ("snippet", "resumen", "nota"))
        for u in extract_urls_from_row(row):
            if not _is_plausible_url(u):
                continue
            conn.execute(
                """
                INSERT OR IGNORE INTO supplier_evidence (supplier_id, batch_id, url, title, snippet, source_sheet)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    sid,
                    batch_id,
                    u,
                    str(tit).strip() if tit else None,
                    str(snip).strip() if snip else None,
                    _SHEET_EVIDENCIAS,
                ),
            )


def _apply_exclusion_tier_override(conn: sqlite3.Connection, batch_id: int) -> None:
    conn.execute(
        """
        UPDATE supplier_priority_snapshot
        SET tier = ?, rank_in_list = NULL, confidence_score = NULL, confidence_label = 'exclusión'
        WHERE supplier_id IN (SELECT id FROM supplier_master WHERE is_exclusion = 1)
          AND batch_id = ?
        """,
        (TIER_EXCLUSION, batch_id),
    )


def import_supplier_workbook(
    conn: sqlite3.Connection,
    xlsx_path: Path,
    *,
    source_filename: str | None = None,
    informe_docx: Path | None = None,
) -> int:
    """Import workbook into SQLite. Returns ``supplier_import_batch.id``."""
    ensure_supplier_tables(conn)
    path = xlsx_path.resolve()
    data = path.read_bytes()
    h = hashlib.sha256(data).hexdigest()
    name = source_filename or path.name
    sheets, _missing = load_workbook_sheets(path)
    row_counts = {k: len(v) for k, v in sheets.items()}
    now = now_iso()

    prior_rows = sheets.get("Prioridades") or []
    cat_json = json.dumps(prior_rows, ensure_ascii=False, default=str)
    resumen_text = _build_batch_resumen_note(sheets.get(_SHEET_RESUMEN) or [], informe_docx)

    batch_id = _insert_import_batch(
        conn,
        source_filename=name,
        file_sha256=h,
        imported_at=now,
        row_counts=row_counts,
        category_priorities_json=cat_json,
        resumen_note=resumen_text,
    )
    company_lookup = build_company_domain_lookup(sheets.get(_SHEET_OP50) or [])

    _ingest_candidate_rows(
        conn, batch_id, company_lookup, now, sheets.get(_SHEET_OP50) or [], TIER_TOP50, _SHEET_OP50
    )
    _ingest_candidate_rows(
        conn, batch_id, company_lookup, now, sheets.get(_SHEET_CONTACTO15) or [], TIER_TOP15, _SHEET_CONTACTO15
    )
    _ingest_candidate_rows(
        conn, batch_id, company_lookup, now, sheets.get(_SHEET_ANEXO) or [], TIER_ANEXO, _SHEET_ANEXO
    )
    _ingest_exclusion_rows(conn, batch_id, company_lookup, now, sheets.get(_SHEET_EXCLUSIONES) or [])
    _ingest_evidence_rows(conn, batch_id, company_lookup, sheets.get(_SHEET_EVIDENCIAS) or [])
    _apply_exclusion_tier_override(conn, batch_id)

    conn.commit()
    return batch_id
